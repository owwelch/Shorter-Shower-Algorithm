from openpyxl import load_workbook
import temp_block

def readIn(loc):
    tempProfile = []
    start_index = 3
    wb = load_workbook(loc)
    sheet = wb.active
    # Expected format for files: Number, date, temperature.
    for i in range(start_index, sheet.max_row+1):
        if sheet.cell(i, 1).value:
            tempProfile.append(temp_block.TempBlock((sheet.cell(i, 3).value), sheet.cell(i,1).value, sheet.cell(i,4).value))

    return tempProfile
